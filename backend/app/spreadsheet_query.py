"""
Análise estruturada de planilhas para o chat: o LLM não conta linhas em texto —
o arquivo original (xlsx/csv) é carregado numa tabela DuckDB em memória e a
agregação é computada deterministicamente (SELECT-only, com timeout e limites).
"""
from __future__ import annotations

import re
import threading
from pathlib import Path
from typing import Any

import duckdb

SUPPORTED_EXTENSIONS = {".xlsx", ".xlsm", ".csv"}

MAX_ROWS_LOADED = 200_000       # linhas carregadas por aba (truncagem reportada no schema)
MAX_RESULT_ROWS = 500           # linhas devolvidas por query (LIMIT defensivo)
SAMPLE_ROWS = 3
QUERY_TIMEOUT_SECONDS = 20.0

# SELECT-only: uma instrução, sem side effects. DuckDB não tem multi-statement em
# execute(), mas bloqueamos por keyword também (defesa em profundidade).
_FORBIDDEN_TOKENS = re.compile(
    r"\b(attach|copy|create|insert|update|delete|drop|alter|install|load|export|import|pragma|set|call|vacuum|checkpoint)\b",
    re.IGNORECASE,
)


class SpreadsheetQueryError(ValueError):
    """Erro de uso (arquivo não suportado, SQL inválido) — mensagem vai ao LLM/usuário."""


def _sanitize_identifier(name: str, fallback: str) -> str:
    cleaned = re.sub(r"[^0-9a-zA-Z_]+", "_", (name or "").strip()).strip("_").lower()
    if not cleaned:
        return fallback
    if cleaned[0].isdigit():
        cleaned = f"c_{cleaned}"
    return cleaned


def _dedupe(names: list[str]) -> list[str]:
    seen: dict[str, int] = {}
    result = []
    for name in names:
        if name in seen:
            seen[name] += 1
            result.append(f"{name}_{seen[name]}")
        else:
            seen[name] = 0
            result.append(name)
    return result


def _load_xlsx(con: duckdb.DuckDBPyConnection, path: Path) -> list[dict[str, Any]]:
    from openpyxl import load_workbook

    workbook = load_workbook(path, read_only=True, data_only=True)
    tables: list[dict[str, Any]] = []
    try:
        for index, sheet in enumerate(workbook.worksheets):
            rows_iter = sheet.iter_rows(values_only=True)
            try:
                header_row = next(rows_iter)
            except StopIteration:
                continue
            original_columns = [str(c).strip() if c is not None else f"col_{i+1}" for i, c in enumerate(header_row)]
            columns = _dedupe([
                _sanitize_identifier(name, f"col_{i+1}") for i, name in enumerate(original_columns)
            ])
            if not columns:
                continue
            table_name = _dedupe([_sanitize_identifier(sheet.title, f"sheet_{index+1}")] )[0]
            rows: list[tuple] = []
            truncated = False
            for row in rows_iter:
                if len(rows) >= MAX_ROWS_LOADED:
                    truncated = True
                    break
                normalized = tuple(
                    (str(v) if not isinstance(v, (int, float, bool)) and v is not None else v)
                    for v in (row[: len(columns)] + (None,) * max(0, len(columns) - len(row)))
                )
                rows.append(normalized)
            column_defs = ", ".join(f'"{c}" VARCHAR' for c in columns)
            con.execute(f'CREATE TABLE "{table_name}" ({column_defs})')
            if rows:
                placeholders = ", ".join("?" for _ in columns)
                con.executemany(f'INSERT INTO "{table_name}" VALUES ({placeholders})', rows)
            tables.append(
                {
                    "table": table_name,
                    "sheet": sheet.title,
                    "columns": columns,
                    "original_columns": original_columns,
                    "row_count": len(rows),
                    "truncated": truncated,
                }
            )
    finally:
        workbook.close()
    return tables


def _load_csv(con: duckdb.DuckDBPyConnection, path: Path) -> list[dict[str, Any]]:
    con.execute(
        "CREATE TABLE data AS SELECT * FROM read_csv_auto(?, sample_size=-1) LIMIT ?",
        [str(path), MAX_ROWS_LOADED],
    )
    row_count = con.execute('SELECT COUNT(*) FROM "data"').fetchone()[0]
    columns = [r[0] for r in con.execute("DESCRIBE data").fetchall()]
    return [
        {
            "table": "data",
            "sheet": path.name,
            "columns": columns,
            "original_columns": columns,
            "row_count": row_count,
            "truncated": row_count >= MAX_ROWS_LOADED,
        }
    ]


def open_spreadsheet(path: Path) -> tuple[duckdb.DuckDBPyConnection, list[dict[str, Any]]]:
    ext = path.suffix.lower()
    if ext not in SUPPORTED_EXTENSIONS:
        raise SpreadsheetQueryError(
            f"Extensão '{ext}' não suportada para análise estruturada (suportadas: {sorted(SUPPORTED_EXTENSIONS)})"
        )
    con = duckdb.connect(":memory:")
    tables = _load_csv(con, path) if ext == ".csv" else _load_xlsx(con, path)
    if not tables:
        raise SpreadsheetQueryError("Planilha sem abas com dados")
    return con, tables


def get_schema(path: Path) -> dict[str, Any]:
    """Abas, colunas (sanitizadas p/ SQL + originais), tipos amostrados e 3 linhas de exemplo."""
    con, tables = open_spreadsheet(path)
    try:
        for table in tables:
            sample = con.execute(f'SELECT * FROM "{table["table"]}" LIMIT {SAMPLE_ROWS}').fetchall()
            table["sample_rows"] = [list(row) for row in sample]
        return {"file": path.name, "tables": tables}
    finally:
        con.close()


def validate_sql(sql: str) -> str:
    statement = (sql or "").strip().rstrip(";").strip()
    if not statement:
        raise SpreadsheetQueryError("SQL vazio")
    if ";" in statement:
        raise SpreadsheetQueryError("Apenas uma instrução SQL por consulta")
    lowered = statement.lower()
    if not (lowered.startswith("select") or lowered.startswith("with")):
        raise SpreadsheetQueryError("Apenas consultas SELECT são permitidas")
    match = _FORBIDDEN_TOKENS.search(statement)
    if match:
        raise SpreadsheetQueryError(f"Instrução não permitida na consulta: {match.group(0).upper()}")
    return statement


def run_query(path: Path, sql: str) -> dict[str, Any]:
    """Executa SELECT com timeout; resultado limitado a MAX_RESULT_ROWS linhas."""
    statement = validate_sql(sql)
    con, tables = open_spreadsheet(path)
    try:
        # Interrupção por timeout: DuckDB expõe con.interrupt() thread-safe.
        timer = threading.Timer(QUERY_TIMEOUT_SECONDS, con.interrupt)
        timer.start()
        try:
            cursor = con.execute(statement)
            rows = cursor.fetchmany(MAX_RESULT_ROWS + 1)
            columns = [d[0] for d in cursor.description or []]
        except duckdb.InterruptException as exc:
            raise SpreadsheetQueryError(
                f"Consulta excedeu o tempo limite de {QUERY_TIMEOUT_SECONDS:.0f}s"
            ) from exc
        except duckdb.Error as exc:
            raise SpreadsheetQueryError(f"Erro na consulta SQL: {exc}") from exc
        finally:
            timer.cancel()
        truncated = len(rows) > MAX_RESULT_ROWS
        rows = rows[:MAX_RESULT_ROWS]
        return {
            "file": path.name,
            "tables": [t["table"] for t in tables],
            "columns": columns,
            "rows": [list(row) for row in rows],
            "row_count": len(rows),
            "truncated": truncated,
        }
    finally:
        con.close()
